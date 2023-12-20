"""Generate vto model output report"""
import os
import argparse
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter
from google.cloud import storage

BUCKET_NAME = 'mrdarshan-public'

argParser = argparse.ArgumentParser(description="Upload directories to Google Cloud Storage")
argParser.add_argument("vto_model_output_dir", help="Local VTO model output directory path")
argParser.add_argument("report_output_dir", help="Local CSV Report Output directory path")

args = argParser.parse_args()
local_vto_model_output_dir = args.vto_model_output_dir
local_report_output_dir = args.report_output_dir


def upload_images_to_gcs(vto_model_output_dir):
    """Upload files to GCS using gsutil command."""
    gcs_vto_model_output_dir = f"vto_model_output_{os.path.basename(vto_model_output_dir)}"
    print(f"Uploading to GCS dir: {gcs_vto_model_output_dir}")

    # gsutil command to upload files in parallel
    gsutil_command = f"gsutil -m cp -r {vto_model_output_dir}/* gs://{BUCKET_NAME}/{gcs_vto_model_output_dir}"
    subprocess.run(gsutil_command, shell=True, check=True)

    return gcs_vto_model_output_dir


def generate_excel_report(gcs_input_dir, report_output_dir):
    # Initialize a Google Cloud Storage client
    storage_client = storage.Client()
    input_bucket = storage_client.bucket(BUCKET_NAME)

    # List blobs (files) in the input directory
    blobs = input_bucket.list_blobs(prefix=gcs_input_dir)

    directories = list({blob.name[len(gcs_input_dir):].split('/')[1] for blob in blobs})
    print("directories: ", directories)

    # Write the header to the Excel file
    header = ['Pallu_URL', 'Blouse_URL', 'Body_URL', 'Reconstructed_Image_URL', 'Output_Image_URL']
    excel_output_file = os.path.join(report_output_dir, 'report.xlsx')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header)

    # Iterate through each blob and write file paths to Excel
    print("Generating report...")
    for pattern in directories:
        pallu_path = f'https://storage.googleapis.com/{BUCKET_NAME}/{gcs_input_dir}/{pattern}/pallu.png'
        blouse_path = f'https://storage.googleapis.com/{BUCKET_NAME}/{gcs_input_dir}/{pattern}/blouse.png'
        body_path = f'https://storage.googleapis.com/{BUCKET_NAME}/{gcs_input_dir}/{pattern}/body.png'
        reconstructed_image_path = f'https://storage.googleapis.com/{BUCKET_NAME}/{gcs_input_dir}/{pattern}/reconstructed_image.png'
        output_image_path = f'https://storage.googleapis.com/{BUCKET_NAME}/{gcs_input_dir}/{pattern}/{pattern}.png'

        # Append the file paths to the Excel file
        sheet.append([pallu_path, blouse_path, body_path, reconstructed_image_path, output_image_path])

    # Add additional columns with IMAGE formulas
    num_columns = len(header)
    num_rows = len(directories) + 1  # Including the header row
    for col in range(num_columns):
        for row in range(2, num_rows + 2):  # Assuming data starts from row 2
            header_value = header[col]
            formula = f'=IMAGE({get_column_letter(col + 1)}{row})'
            sheet.cell(row=row, column=num_columns + col + 1, value=formula)
            sheet.cell(row=1, column=num_columns + col + 1, value=header_value.replace('_URL', '_IMG'))

    # Save the workbook to a file
    workbook.save(excel_output_file)

    print(f"Excel file '{excel_output_file}' created successfully.")


if __name__ == "__main__":
    # Upload images to gcs bucket
    gcs_vto_model_output_dir = upload_images_to_gcs(vto_model_output_dir=local_vto_model_output_dir)

    # Generate report
    generate_excel_report(gcs_vto_model_output_dir, local_report_output_dir)
